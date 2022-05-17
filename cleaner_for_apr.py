from openpyxl import Workbook, load_workbook
from fuzzywuzzy import process, fuzz
import time

path_to_docs = 'C:/Users/PetukhovMD/Desktop/справочники/'
path_to_data = 'C:/Users/PetukhovMD/Desktop/2021/'

wb = load_workbook(path_to_docs + 'input.xlsx')
groups = {}
trans = {}
ws = wb['должности']

for row in ws.rows:
    if row[0].value == 'Должности НЕ вкл. в фильтр':
        continue
    trans[row[0].value] = row[1].value

ws = wb['УГД']
look_up = list(trans.keys())
for row in ws.rows:
    if row[0].value == 'Должность':
        continue
    groups[row[0].value] = [row[1].value, row[2].value, row[3].value]

ws = load_workbook(path_to_docs + 'ИНН+ЕКИС.xlsx').active
inn_by_school = {}

for row in ws.rows:
    if row[1].value in ['Код ЕКИС', None] :
        continue
    inn_by_school[row[6].value] = [row[3].value, row[7].value]

def timeit(a, *kargs, **kwargs):
    def timed_a(*kargs, **kwargs):
        t = time.time()
        a(*kargs, **kwargs)
        print(time.time() - t)
    return timed_a

@timeit
def calc():
    base = load_workbook(path_to_data + 'апрель.xlsx').active
    wb = Workbook()
    ws = wb.active
    ws.append(["Наименование организации","СНИЛС","Всего начислено с учетом распределения отпуска","Количество ставок на начало месяца","type", 'status', "Должность по 106р","day"])

    #i = 0

    for row in base.rows:
        # i += 1
        # if i > 10:
        #     break
        if row[2].value in [None, 'ФИО']:
            continue
        res = []
        res.append(row[0].value)
        res.append(row[1].value)
        res.append(row[4].value)
        res.append(row[6].value)
        res.append(row[7].value)
        res.append(row[9].value)
        res.append(process.extractOne(row[10].value, look_up, scorer=fuzz.WRatio)[0])
        if (row[-1].value == None) | (row[-2].value == None):
            res.append(0)
        else:
            res.append(1 if int(row[-1].value) <= int(row[-2].value) else 0)
        ws.append(res)
    wb.save(path_to_data + 'clean_slow.xlsx')

calc()