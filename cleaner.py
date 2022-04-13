from openpyxl import Workbook, load_workbook
from fuzzywuzzy import process, fuzz
import time

path_to_docs = 'C:/Users/PetukhovMD/Desktop/справочники/'
path_to_data = 'C:/Users/PetukhovMD/Desktop/2022/март/'

wb = load_workbook(path_to_docs + 'input.xlsx')
groups = {}
trans = {}
ws = wb['должности']

for row in ws.rows:
    if row[0].value == 'Должности НЕ вкл. в фильтр':
        continue
    trans[row[0].value] = row[1].value

ws = wb['УГД']
look_up = list(trans.keys())
for row in ws.rows:
    if row[0].value == 'Должность':
        continue
    groups[row[0].value] = [row[1].value, row[2].value, row[3].value]

ws = load_workbook(path_to_docs + 'ИНН+ЕКИС.xlsx').active
inn_by_school = {}

for row in ws.rows:
    if row[1].value in ['Код ЕКИС', None] :
        continue
    inn_by_school[row[6].value] = [row[3].value, row[7].value]

def timeit(a, *kargs, **kwargs):
    def timed_a(*kargs, **kwargs):
        t = time.time()
        a(*kargs, **kwargs)
        print(time.time() - t)
    return timed_a

@timeit
def calc():
    base = load_workbook(path_to_data + 'ДОНМ_Начисления работников.xlsx').active
    wb = Workbook()
    ws = wb.active
    ws.append(["Инн орг",'Тип орг', "Наименование организации","СНИЛС","ФИО","Всего начислено с учетом распределения отпуска","Количество ставок на начало месяца","Бинарная работа","Должность по 106р",'ugd', 'ped', 'neposr isp',"бинарный день"])

    for row in base.rows:
        if row[2].value in [None, 'ОрганизацияИНН']:
            continue
        res = []
        res.append(row[2].value)
        i = int(row[2].value)
        if i in inn_by_school:
            res.append(inn_by_school[i][1])
            res.append(inn_by_school[i][0])
        else:
            res.append('')
            res.append('cant resolve school sname')
        res.append(row[4].value)
        res.append(row[5].value)
        res.append(row[-11].value)
        res.append(row[10].value)
        res.append(1 if row[-9].value == 'Основное место работы' and row[-7].value ==  'Работа' else 0)
        if  not row[8].value:
            res.append('no job value')
            res.append('')
            res.append('')
            res.append('')
        else:
            choice = process.extractOne(row[8].value, look_up, scorer=fuzz.WRatio)[0]
            if choice in trans:
                res.append(trans[choice])
                res.append(groups[trans[choice]][0])
                res.append(groups[trans[choice]][1])
                res.append(groups[trans[choice]][2])
            else:
                res.append('cant resolve job title')
                res.append(row[4].value + "+" + row[5].value + '+' + row[-6].value)
                res.append('')
                res.append('')
                print(row[4].value, "+",row[5].value, '+',row[-6].value)
        if (row[-4].value == None) | (row[-5].value == None):
            res.append(0)
        else:
            res.append(1 if int(row[-4].value) <= int(row[-5].value) else 0)
        ws.append(res)
    wb.save(path_to_data + 'clean_slow.xlsx')

calc()